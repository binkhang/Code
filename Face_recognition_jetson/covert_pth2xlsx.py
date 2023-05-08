import openpyxl
import torch
import os

DATA_PATH = './Face_recognition/encoded_data/minimized_embeddings'
# DATA_PATH = './keypadHandler'
embeddings = torch.load(os.path.join(DATA_PATH, "user_2911.pth"))
print(len(embeddings))

workbook = openpyxl.Workbook()
worksheet = workbook.active

for i, embedding in enumerate(embeddings):
    for j, value in enumerate(embedding):
        cell = worksheet.cell(row=i+1, column=j+1)
        cell.value = value.item()

workbook.save(os.path.join(DATA_PATH, "user_2911.xlsx"))


# DATA_PATH = './Face_recognition/encoded_data/'
 
# embeddings = torch.load(os.path.join(DATA_PATH, "faceslist_from_infer_ipynb.pth"))
# print(len(embeddings))
# embeddings.double()
# workbook = openpyxl.Workbook()
# worksheet = workbook.active

# for i, embedding in enumerate(embeddings):
#     for j, value in enumerate(embedding):
#         cell = worksheet.cell(row=i+1, column=j+1)
#         cell.value = value.item()

# workbook.save(os.path.join(DATA_PATH, "embeddings_from_infer_ipynb.xlsx"))
